"""Source adapters that discover content for indexing."""

from __future__ import annotations

import json
import logging
import mailbox
import re
from datetime import datetime
from email import policy
from email.message import EmailMessage
from email.parser import BytesParser
from pathlib import Path
from typing import Iterator, Protocol

from .models import SourceItem

logger = logging.getLogger(__name__)


class SourceAdapter(Protocol):
    name: str
    def discover(self, root: Path) -> Iterator[SourceItem]: ...


# ---------------------------------------------------------------------------
# FsTextAdapter
# ---------------------------------------------------------------------------

class FsTextAdapter:
    name = "fs-text"

    EXTENSIONS = {".txt", ".md", ".markdown", ".text"}

    def discover(self, root: Path) -> Iterator[SourceItem]:
        if not root.exists():
            return
        if root.is_file() and root.suffix.lower() in self.EXTENSIONS:
            yield from self._process_file(root)
            return
        for path in root.rglob("*"):
            if not path.is_file():
                continue
            if path.suffix.lower() not in self.EXTENSIONS:
                continue
            yield from self._process_file(path)

    def _process_file(self, path: Path) -> Iterator[SourceItem]:
        try:
            body = path.read_text(encoding="utf-8", errors="replace")
        except OSError:
            return
        stat = path.stat()
        yield SourceItem(
            source_ref=str(path),
            title=path.name,
            modified_at=datetime.fromtimestamp(stat.st_mtime),
            body=body,
            meta={"size_bytes": stat.st_size},
        )


# ---------------------------------------------------------------------------
# FsDocAdapter
# ---------------------------------------------------------------------------

class FsDocAdapter:
    name = "fs-doc"

    EXTENSIONS = {".pdf", ".docx", ".xlsx"}

    def discover(self, root: Path) -> Iterator[SourceItem]:
        if not root.exists():
            return
        if root.is_file() and root.suffix.lower() in self.EXTENSIONS:
            yield from self._process_file(root)
            return
        for path in root.rglob("*"):
            if not path.is_file():
                continue
            if path.suffix.lower() not in self.EXTENSIONS:
                continue
            yield from self._process_file(path)

    def _process_file(self, path: Path) -> Iterator[SourceItem]:
        ext = path.suffix.lower()
        body = ""
        try:
            if ext == ".pdf":
                try:
                    from pypdf import PdfReader
                    reader = PdfReader(path)
                    body = "\n\n".join(page.extract_text() or "" for page in reader.pages)
                except ImportError:
                    logger.warning("pypdf not installed. Skipping %s", path)
                    return
            elif ext == ".docx":
                try:
                    from docx import Document
                    doc = Document(path)
                    body = "\n".join(p.text for p in doc.paragraphs)
                except ImportError:
                    logger.warning("python-docx not installed. Skipping %s", path)
                    return
            elif ext == ".xlsx":
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(path, read_only=True, data_only=True)
                    rows = []
                    for sheet in wb.worksheets:
                        for row in sheet.iter_rows(values_only=True):
                            row_vals = [str(cell) for cell in row if cell is not None]
                            if row_vals:
                                rows.append(" | ".join(row_vals))
                    body = "\n".join(rows)
                except ImportError:
                    logger.warning("openpyxl not installed. Skipping %s", path)
                    return

            stat = path.stat()
            yield SourceItem(
                source_ref=str(path),
                title=path.name,
                modified_at=datetime.fromtimestamp(stat.st_mtime),
                body=body.strip(),
                meta={"size_bytes": stat.st_size},
            )
        except Exception as e:
            logger.debug("Failed to parse %s: %s", path, e)


# ---------------------------------------------------------------------------
# MailAdapter
# ---------------------------------------------------------------------------

class MailAdapter:
    name = "mail"

    def discover(self, root: Path) -> Iterator[SourceItem]:
        if not root.exists():
            return
        if root.is_file():
            yield from self._process_path(root)
            return
        for path in root.rglob("*"):
            if not path.is_file():
                continue
            yield from self._process_path(path)

    def _process_path(self, path: Path) -> Iterator[SourceItem]:
        ext = path.suffix.lower()
        if ext == ".mbox":
            try:
                mbox = mailbox.mbox(path)
                for i, msg in enumerate(mbox):
                    yield self._msg_to_item(msg, f"{path}:{i}")
            except Exception as e:
                logger.debug("Failed to read mbox %s: %s", path, e)
        elif ext == ".eml":
            try:
                with open(path, "rb") as f:
                    msg = BytesParser(policy=policy.default).parse(f)
                yield self._msg_to_item(msg, str(path))
            except Exception as e:
                logger.debug("Failed to read eml %s: %s", path, e)

    def _msg_to_item(self, msg: mailbox.Message | EmailMessage, ref: str) -> SourceItem:
        subject = str(msg.get("subject", "(no subject)"))
        date_str = str(msg.get("date", ""))
        from_str = str(msg.get("from", ""))
        to_str = str(msg.get("to", ""))

        body = ""
        if msg.is_multipart():
            for part in msg.walk():
                ctype = part.get_content_type()
                cdispo = str(part.get("Content-Disposition", ""))
                if ctype == "text/plain" and "attachment" not in cdispo:
                    payload = part.get_payload(decode=True)
                    if payload:
                        body += payload.decode("utf-8", errors="replace") + "\n"
        else:
            payload = msg.get_payload(decode=True)
            if payload:
                body = payload.decode("utf-8", errors="replace")

        return SourceItem(
            source_ref=ref,
            title=subject,
            body=body.strip() or "(empty body)",
            meta={"from": from_str, "to": to_str, "date": date_str},
        )


# ---------------------------------------------------------------------------
# SignalExportAdapter
# ---------------------------------------------------------------------------

class SignalExportAdapter:
    name = "signal-export"

    def discover(self, root: Path) -> Iterator[SourceItem]:
        if not root.exists():
            return
        if root.is_file() and root.suffix.lower() == ".json":
            yield from self._process_file(root)
            return
        for path in root.rglob("*.json"):
            if path.is_file():
                yield from self._process_file(path)

    def _process_file(self, path: Path) -> Iterator[SourceItem]:
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            
            # Assume data is a list of threads/conversations
            if isinstance(data, list):
                for conv_idx, conv in enumerate(data):
                    title = conv.get("name") or conv.get("title") or f"Conversation {conv_idx}"
                    messages = conv.get("messages", [])
                    for i, msg in enumerate(messages):
                        body = msg.get("body")
                        if not body:
                            continue
                        yield SourceItem(
                            source_ref=f"{path}:{conv_idx}:{i}",
                            title=title,
                            body=body,
                            meta={
                                "timestamp": msg.get("timestamp"),
                                "type": msg.get("type"),
                                "sender": msg.get("source") or msg.get("sender"),
                            }
                        )
        except Exception as e:
            logger.debug("Failed to parse Signal export %s: %s", path, e)


# ---------------------------------------------------------------------------
# WhatsAppTextAdapter
# ---------------------------------------------------------------------------

class WhatsAppTextAdapter:
    name = "whatsapp-text"

    # Regex for WhatsApp lines: "1/15/23, 10:30 AM - Sender Name: Message body"
    # Matches various date formats globally.
    MSG_PATTERN = re.compile(r'^(\d{1,2}[/-]\d{1,2}[/-]\d{2,4},\s*\d{1,2}:\d{2}(?:\s*[AP]M)?)\s*-\s*(.+?):\s*(.+)$')

    def discover(self, root: Path) -> Iterator[SourceItem]:
        if not root.exists():
            return
        if root.is_file() and root.suffix.lower() == ".txt":
            yield from self._process_file(root)
            return
        for path in root.rglob("*.txt"):
            if path.is_file():
                yield from self._process_file(path)

    def _process_file(self, path: Path) -> Iterator[SourceItem]:
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
        except OSError:
            return

        lines = text.splitlines()
        current_sender = None
        current_date = None
        current_body = []
        start_line = 0

        def emit_block(end_line: int) -> SourceItem | None:
            if current_body and current_sender:
                return SourceItem(
                    source_ref=f"{path}:{start_line}-{end_line}",
                    title=f"WhatsApp chat with {current_sender}",
                    body="\n".join(current_body),
                    meta={"sender": current_sender, "date": current_date}
                )
            return None

        for i, line in enumerate(lines, start=1):
            m = self.MSG_PATTERN.match(line)
            if m:
                # Flush previous block
                if current_sender and current_body:
                    if item := emit_block(i - 1):
                        yield item
                
                date_str, sender, msg = m.groups()
                current_date = date_str
                current_sender = sender
                current_body = [msg]
                start_line = i
            else:
                # Continuation of previous message or system message
                if current_body:
                    current_body.append(line)

        # Flush the final block
        if item := emit_block(len(lines)):
            yield item


# ---------------------------------------------------------------------------
# PhotosExifAdapter
# ---------------------------------------------------------------------------

class PhotosExifAdapter:
    name = "photos-exif"
    EXTENSIONS = {".jpg", ".jpeg", ".png", ".tiff", ".heic"}

    def discover(self, root: Path) -> Iterator[SourceItem]:
        if not root.exists():
            return
        if root.is_file() and root.suffix.lower() in self.EXTENSIONS:
            yield from self._process_file(root)
            return
        for path in root.rglob("*"):
            if not path.is_file():
                continue
            if path.suffix.lower() not in self.EXTENSIONS:
                continue
            yield from self._process_file(path)

    def _process_file(self, path: Path) -> Iterator[SourceItem]:
        try:
            from PIL import Image, ExifTags
        except ImportError:
            logger.warning("Pillow not installed. Skipping photo exif %s", path)
            return

        try:
            with Image.open(path) as img:
                exif = img.getexif()
                meta = {}
                if exif:
                    for tag_id, val in exif.items():
                        tag = ExifTags.TAGS.get(tag_id, tag_id)
                        if isinstance(val, (bytes, int, str, float)):
                            meta[str(tag)] = str(val)

                body_lines = [f"Photo: {path.name}"]
                if meta:
                    body_lines.append("EXIF Data:")
                    for k, v in meta.items():
                        body_lines.append(f"  {k}: {v}")
                
                stat = path.stat()
                meta["size_bytes"] = stat.st_size

                yield SourceItem(
                    source_ref=str(path),
                    title=path.name,
                    modified_at=datetime.fromtimestamp(stat.st_mtime),
                    body="\n".join(body_lines),
                    meta=meta
                )
        except Exception as e:
            logger.debug("Failed to process image %s: %s", path, e)


REGISTRY: dict[str, SourceAdapter] = {
    FsTextAdapter.name: FsTextAdapter(),
    FsDocAdapter.name: FsDocAdapter(),
    MailAdapter.name: MailAdapter(),
    SignalExportAdapter.name: SignalExportAdapter(),
    WhatsAppTextAdapter.name: WhatsAppTextAdapter(),
    PhotosExifAdapter.name: PhotosExifAdapter(),
}


def get_adapter(name: str) -> SourceAdapter:
    if name not in REGISTRY:
        raise KeyError(f"Unknown adapter: {name}. Available: {', '.join(REGISTRY)}")
    return REGISTRY[name]
